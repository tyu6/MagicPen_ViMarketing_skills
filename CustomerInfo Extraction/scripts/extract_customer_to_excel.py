#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit('缺少依赖 openpyxl，请先执行: pip install openpyxl') from exc

TEXT_SUFFIXES={'.md','.txt','.log','.csv','.json','.yaml','.yml','.srt','.text',''}
ENCODINGS=['utf-8-sig','utf-8','gb18030','gbk','gb2312','utf-16','big5','latin-1']
GENERIC_LABELS={'客户','用户','对方','采购','甲方','联系人','访客'}
SYSTEM_PATTERNS=[r'撤回了一条消息',r'加入了群聊',r'退出了群聊',r'系统消息',r'消息已发出',r'以上为.*聊天记录',r'^图片$',r'^语音$',r'^视频$',r'^文件$',r'^\[表情\]$']
JOB_TITLES=['CEO','CTO','COO','VP','总经理','副总','采购经理','总监','经理','负责人','采购','老板','主任','主管','顾问','工程师','运营','销售','客服','老师']
PROVINCES=['北京','上海','天津','重庆','河北','山西','辽宁','吉林','黑龙江','江苏','浙江','安徽','福建','江西','山东','河南','湖北','湖南','广东','海南','四川','贵州','云南','陕西','甘肃','青海','内蒙古','广西','西藏','宁夏','新疆','香港','澳门']
CITIES=['北京','上海','深圳','广州','杭州','苏州','南京','武汉','成都','重庆','西安','长沙','郑州','青岛','厦门','福州','宁波','东莞','佛山','合肥','济南','天津']
INDUSTRY_KEYWORDS={'制造':['制造','工厂','产线','设备'],'零售':['零售','门店','电商','商超'],'金融':['银行','保险','证券','基金'],'医疗':['医院','医疗','药','诊所'],'教育':['学校','教育','培训','教培'],'互联网':['SaaS','互联网','平台','APP','软件'],'物流':['物流','仓储','配送','运单'],'地产':['地产','物业','楼盘','房产']}
PRODUCT_KEYWORDS={'客服质检':['质检','会话质检','客服质检'],'工单系统':['工单','派单','ticket'],'CRM':['CRM','客户管理'],'营销自动化':['营销','自动化','私域','线索培育'],'智能客服':['机器人','智能客服','客服机器人'],'外呼':['外呼','电销','呼叫中心'],'数据分析':['分析','报表','BI','数据看板']}
SOURCE_PATTERNS={'官网':[r'官网',r'官网留资'],'转介绍':[r'转介绍',r'朋友介绍',r'同事介绍'],'活动':[r'活动',r'展会',r'峰会',r'沙龙'],'广告':[r'广告',r'投放',r'信息流'],'私域':[r'私域',r'朋友圈',r'社群'],'陌拜':[r'陌拜',r'cold call',r'陌生开发']}
STATUS_RULES=[('已搁置',[r'暂时不考虑',r'先不推进',r'以后再说',r'没预算',r'今年不做',r'已搁置']),('报价中',[r'报价',r'合同',r'招标',r'付款',r'采购流程']),('方案评估',[r'评估',r'比较',r'PoC',r'POC',r'试用',r'demo',r'演示']),('需求沟通',[r'需求',r'场景',r'对接',r'集成',r'想了解',r'想看看'])]
KEY_QUOTE_PATTERNS=[r'预算',r'报价',r'需求',r'场景',r'问题',r'上线',r'试用',r'PoC',r'对接',r'公司',r'微信',r'邮箱']

def load_field_config():
    path=Path(__file__).resolve().parent.parent/'config'/'field_mapping.json'
    return json.loads(path.read_text(encoding='utf-8'))

def read_text_file(path:Path):
    raw=path.read_bytes()
    if not raw:
        raise SystemExit(f'输入文件为空: {path}')
    if raw.count(b'\x00')/max(len(raw),1)>0.05:
        raise SystemExit(f'文件看起来像二进制，无法按文本处理: {path}')
    for enc in ENCODINGS:
        try:
            return raw.decode(enc),enc
        except UnicodeDecodeError:
            pass
    raise SystemExit(f'无法用常见编码解析文件: {path}')

def normalize_text(text:str):
    text=text.replace('\r\n','\n').replace('\r','\n').replace('\u3000',' ')
    return re.sub(r'[ \t]+',' ',text).strip()

def detect_timestamp(text:str):
    for p in [r'(20\d{2}[-/]\d{1,2}[-/]\d{1,2}\s+\d{1,2}:\d{2}(?::\d{2})?)',r'(20\d{2}[-/]\d{1,2}[-/]\d{1,2})',r'(\d{1,2}:\d{2}(?::\d{2})?)']:
        m=re.search(p,text)
        if m:
            return m.group(1)
    return ''

def guess_speaker(label:str):
    low=label.lower()
    if any(t in low for t in ['销售','商务','顾问','客服','售前','运营','bd','ae','我方']):
        return 'seller'
    if any(t in low for t in ['客户','甲方','采购','用户','对方','客户方','负责人','老板','总','经理']):
        return 'customer'
    return 'unknown'

def parse_lines(text:str):
    lines=[]
    sys_res=[re.compile(p,re.I) for p in SYSTEM_PATTERNS]
    for idx,raw in enumerate(text.split('\n'),start=1):
        raw=raw.strip(); raw=re.sub(r'^>+\s*','',raw); raw=re.sub(r'^\[[^\]]+\]\s*','',raw)
        if not raw or any(r.search(raw) for r in sys_res):
            continue
        ts=detect_timestamp(raw)
        stripped=re.sub(r'^20\d{2}[-/]\d{1,2}[-/]\d{1,2}\s+\d{1,2}:\d{2}(?::\d{2})?\s*','',raw)
        stripped=re.sub(r'^\d{1,2}:\d{2}(?::\d{2})?\s*','',stripped)
        label=''; speaker='unknown'; content=stripped
        m=re.match(r'^(?P<label>[^:：]{1,32})[:：]\s*(?P<content>.+)$',stripped)
        if m:
            label=m.group('label').strip(); content=m.group('content').strip(); speaker=guess_speaker(label)
        lines.append({'index':idx,'raw':raw,'content':content,'speaker_label':label,'speaker':speaker,'timestamp':ts})
    return lines

def normalize_list(values):
    out=[]
    for v in values:
        v=v.strip().strip('，,。;；')
        if v and v not in out:
            out.append(v)
    return out

def add_candidate(store,field,value,line_index,weight=1):
    value=value.strip()
    if value:
        store[field].append({'value':value,'line_index':line_index,'weight':weight})

def pick_best(candidates):
    if not candidates:
        return '',[]
    ordered=sorted(candidates,key=lambda x:(x['weight'],x['line_index'],len(x['value'])),reverse=True)
    best=ordered[0]['value']
    conflicts=normalize_list([x['value'] for x in ordered if x['value']!=best])
    return best,conflicts

def looks_like_real_name(value:str):
    if not value or value in GENERIC_LABELS or re.fullmatch(r'[A-Za-z][A-Za-z0-9_\-]{1,20}',value):
        return False
    pure=re.sub(r'(总|经理|老师|主任|总监|采购|先生|女士)$','',value)
    return bool(re.fullmatch(r'[\u4e00-\u9fa5]{2,4}',pure))

def normalize_name(value:str):
    value=re.sub(r'^(客户|用户|采购|联系人|客户方)[-_]','',value.strip())
    value=re.sub(r'^(我是|我叫|叫我)','',value)
    return value.strip(' ：:-')

def extract_company_names(text:str):
    pats=[r'([\u4e00-\u9fa5A-Za-z0-9（）()·&]{2,30}(?:有限公司|有限责任公司|集团|科技|信息|网络|银行|医院|学校|研究院|事务所|制造|电子|医药|公司))(?:的|，|,|。|$)']
    out=[]
    for pat in pats:
        for m in re.findall(pat,text):
            v=(m if isinstance(m,str) else m[0]).strip(' ，,。')
            v=re.sub(r'^(我是|来自|这边是|我在)','',v)
            if len(v)>=3 and v not in out:
                out.append(v)
    return out

def company_short_name(name:str):
    for suf in ['有限责任公司','股份有限公司','有限公司','集团','科技','信息','网络','制造','电子','医药','公司']:
        if name.endswith(suf) and len(name)>len(suf):
            return name[:-len(suf)].strip('（）()')
    return name.strip('（）()')

def infer_industry(text:str):
    for industry,keys in INDUSTRY_KEYWORDS.items():
        if any(k.lower() in text.lower() for k in keys):
            return industry
    return ''

def infer_source_channel(text:str):
    for source,pats in SOURCE_PATTERNS.items():
        if any(re.search(p,text,re.I) for p in pats):
            return source
    return ''

def infer_product_interest(text:str):
    hits=[]
    for product,keys in PRODUCT_KEYWORDS.items():
        if any(k.lower() in text.lower() for k in keys):
            hits.append(product)
    return '、'.join(hits)

def extract_first(pattern:str,text:str):
    m=re.search(pattern,text,re.I)
    return m.group(1).strip() if m else ''

def extract_lines_by_patterns(lines,patterns,max_items=3):
    out=[]; regs=[re.compile(p,re.I) for p in patterns]
    for line in lines:
        if any(r.search(line['content']) for r in regs):
            v=line['content'].strip()
            if v not in out:
                out.append(v)
        if len(out)>=max_items:
            break
    return out

def format_list_value(items):
    return '；'.join(normalize_list(items))

def infer_status(text:str):
    for status,pats in STATUS_RULES:
        if any(re.search(p,text,re.I) for p in pats):
            return status
    return '初步接触'

def infer_customer_type(text:str,company_name_value:str):
    if re.search(r'代理|渠道|经销',text):
        return '渠道'
    if company_name_value:
        return '企业'
    if re.search(r'个人|自己用|个人使用',text):
        return '个人'
    if re.search(r'合作伙伴|代理商',text):
        return '代理'
    return '其它'

def infer_opportunity_level(status,budget_info,timeline_info,decision_maker,risk_flags,contact_strength):
    if any(flag == '信息不足' or '搁置' in flag or '没预算' in flag or '拒绝' in flag for flag in risk_flags):
        return 'low'
    if status=='报价中':
        return 'high'
    if status=='方案评估' and budget_info and timeline_info and contact_strength:
        return 'high'
    if status in {'方案评估','需求沟通'} and contact_strength:
        return 'medium'
    if decision_maker and (budget_info or timeline_info):
        return 'medium'
    return 'low'

def infer_next_step(status,missing_fields,risk_flags):
    if status=='已搁置':
        return '记录搁置原因，降低跟进频率，等待客户重新出现窗口。'
    if status=='报价中':
        return '跟进报价反馈，确认决策人、采购流程和最终时间表。'
    if status=='方案评估':
        return '安排演示或 PoC，补齐预算、时间表和技术对接人。'
    if status=='需求沟通':
        if any(f in missing_fields for f in ['customer_name','company_name','mobile','email','wechat']):
            return '优先补齐联系人和联系方式，再推动一次短会明确需求范围。'
        return '继续澄清场景和痛点，推动形成明确方案或报价动作。'
    if '信息不足' in risk_flags:
        return '先补充更多对话内容或关键身份信息，再继续抽取和跟进。'
    return '先确认客户身份、需求和下一次沟通动作。'

def extract_region(text:str):
    province=next((i for i in PROVINCES if i in text),'')
    city=next((i for i in CITIES if i in text),'')
    if city in {'北京','上海','天津','重庆'} and not province:
        province=city
    return city,province

def suspicious_flags(record):
    flags=[]
    if record['mobile'] and not re.fullmatch(r'1[3-9]\d{9}',record['mobile']):
        flags.append('手机号格式可疑')
    if record['phone'] and not re.fullmatch(r'(?:0\d{2,3}-?)?\d{7,8}',record['phone']):
        flags.append('固话格式可疑')
    if record['email'] and not re.fullmatch(r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}',record['email']):
        flags.append('邮箱格式可疑')
    if record['company_name'] and len(record['company_name'])<4:
        flags.append('公司名过短，可能是昵称')
    return normalize_list(flags)

def split_groups(lines):
    label_keys=[]
    for line in lines:
        label=line['speaker_label']
        if line['speaker'] in {'customer','unknown'} and label and label not in GENERIC_LABELS:
            norm=normalize_name(label)
            if looks_like_real_name(norm) and norm not in label_keys:
                label_keys.append(norm)
    if len(label_keys)<=1:
        return [lines],''
    groups=[[]]; current_key=''; seen_keys=set(); note=''
    for line in lines:
        key=''; label=line['speaker_label']
        if line['speaker'] in {'customer','unknown'} and label and label not in GENERIC_LABELS:
            norm=normalize_name(label)
            if looks_like_real_name(norm):
                key=norm
        key=extract_first(r'(1[3-9]\d{9})',line['content']) or extract_first(r'([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})',line['content']) or key
        if key and current_key and key!=current_key and groups[-1]:
            seen_keys.add(current_key)
            if key not in seen_keys:
                groups.append([])
        if key:
            current_key=key
        groups[-1].append(line)
    groups=[g for g in groups if g]
    if 1<len(groups)<=5:
        note=f'检测到 {len(groups)} 个可能客户分组，已按显式身份线索拆分。'
        return groups,note
    if len(groups)>5:
        note='文件中疑似出现多个客户，但无法可靠拆分，已合并输出。'
    return [lines],note

def build_blank_record(source_file,extraction_time,raw_text_length):
    return {'customer_name':'','gender_guess':'','phone':'','mobile':'','email':'','wechat':'','company_name':'','company_short_name':'','department':'','job_title':'','city':'','province':'','industry':'','customer_type':'其它','source_channel':'','product_interest':'','use_case':'','pain_points':'','budget_info':'','timeline_info':'','decision_maker':'','current_status':'初步接触','next_step':'先提供更多可读文本内容。','opportunity_level':'low','confidence':10,'last_contact_time':'','communication_summary':'未识别到有效对话内容。','key_quotes':[],'risk_flags':['信息不足'],'remarks':'文件内容为空或均为噪声。','source_file':source_file,'extraction_time':extraction_time,'raw_text_length':raw_text_length,'missing_fields':['customer_name','company_name','mobile','email','wechat']}

def extract_record(lines,source_file,extraction_time,raw_text_length,group_note):
    cands=defaultdict(list); conflict_notes=[]; quotes=[]; customer_lines=[]; timestamps=[]; text_blob='\n'.join(l['content'] for l in lines)
    for line in lines:
        content=line['content']
        if line['timestamp']:
            timestamps.append(line['timestamp'])
        if line['speaker'] in {'customer','unknown'}:
            customer_lines.append(content)
        for mobile in re.findall(r'(?<!\d)(1[3-9]\d{9})(?!\d)',content):
            add_candidate(cands,'mobile',mobile,line['index'],5)
        for phone in re.findall(r'(?<!\d)((?:0\d{2,3}-?)?\d{7,8})(?!\d)',content):
            if not re.fullmatch(r'1[3-9]\d{9}',phone):
                add_candidate(cands,'phone',phone,line['index'],4)
        for email in re.findall(r'([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})',content):
            add_candidate(cands,'email',email,line['index'],5)
        wechat=extract_first(r'(?:微信|微信号|vx|wx)(?:是|号)?[:： ]*([A-Za-z][A-Za-z0-9_-]{5,20})',content)
        if wechat:
            add_candidate(cands,'wechat',wechat,line['index'],5)
        if line['speaker_label'] and line['speaker']=='customer':
            label_name=normalize_name(line['speaker_label'])
            if looks_like_real_name(label_name):
                clean_name=re.sub(r'(先生|女士|经理|总监|主任|采购)$','',label_name)
                add_candidate(cands,'customer_name',clean_name,line['index'],10)
            else:
                add_candidate(cands,'remarks_name_hint',label_name,line['index'],1)
        explicit_name=extract_first(r'(?:我叫|姓名[:：]?)([\u4e00-\u9fa5]{2,4})',content) or extract_first(r'(?:采购经理|经理|总监|主任|主管|顾问)([\u4e00-\u9fa5]{2,4})(?:[，,。]|$)',content)
        if explicit_name:
            add_candidate(cands,'customer_name',explicit_name,line['index'],6)
        for company in extract_company_names(content):
            add_candidate(cands,'company_name',company,line['index'],5 if re.search(r'公司|集团|科技|制造|银行|医院',company) else 2)
        dept=extract_first(r'([\u4e00-\u9fa5A-Za-z0-9]{2,20}(?:部|中心|事业部|办公室|团队))',content)
        if dept:
            add_candidate(cands,'department',dept,line['index'],3)
        for title in JOB_TITLES:
            title_weight=len(title)
            content_hit=title.lower() in content.lower() and title not in {'销售','客服'}
            label_hit=title.lower() in line['speaker_label'].lower()
            if title == '负责人' and '技术负责人' in content:
                continue
            if line['speaker'] != 'seller' and (content_hit or label_hit):
                add_candidate(cands,'job_title',title,line['index'],title_weight)
                break
        if re.search(r'女士|小姐|Ms\.?',content,re.I):
            add_candidate(cands,'gender_guess','female',line['index'],2)
        elif re.search(r'先生|Mr\.?',content,re.I):
            add_candidate(cands,'gender_guess','male',line['index'],2)
        city,province=extract_region(line['raw'])
        region_weight=5 if ('我是' in content or '公司' in content or line['index'] == 1) else 1
        if city:
            add_candidate(cands,'city',city,line['index'],region_weight)
        if province:
            add_candidate(cands,'province',province,line['index'],region_weight)
        if any(re.search(p,content,re.I) for p in KEY_QUOTE_PATTERNS):
            quotes.append(content)
    customer_text='\n'.join(customer_lines) if customer_lines else text_blob
    status=infer_status(text_blob)
    record={'customer_name':'','gender_guess':'','phone':'','mobile':'','email':'','wechat':'','company_name':'','company_short_name':'','department':'','job_title':'','city':'','province':'','industry':infer_industry(text_blob) or infer_industry(customer_text),'customer_type':'','source_channel':infer_source_channel(text_blob) or infer_source_channel(customer_text),'product_interest':infer_product_interest(text_blob) or infer_product_interest(customer_text),'use_case':format_list_value(extract_lines_by_patterns(lines,[r'用于',r'场景',r'对接',r'集成',r'想做',r'需要',r'系统'],2)),'pain_points':format_list_value(extract_lines_by_patterns(lines,[r'痛点',r'问题',r'困难',r'低效',r'成本高',r'跟进慢',r'人工'],2)),'budget_info':format_list_value(extract_lines_by_patterns(lines,[r'预算',r'报价',r'价格',r'多少钱',r'费用',r'万',r'元'],2)),'timeline_info':format_list_value(extract_lines_by_patterns(lines,[r'这周',r'下周',r'本月',r'下月',r'月底',r'尽快',r'上线',r'排期',r'时间'],2)),'decision_maker':format_list_value(extract_lines_by_patterns(lines,[r'决策',r'老板',r'采购',r'法务',r'审批',r'拍板',r'我能定'],2)),'current_status':status,'next_step':'','opportunity_level':'','confidence':0,'last_contact_time':timestamps[-1] if timestamps else '','communication_summary':'','key_quotes':normalize_list(quotes)[:5],'risk_flags':[],'remarks':'','source_file':source_file,'extraction_time':extraction_time,'raw_text_length':raw_text_length,'missing_fields':[]}
    for field in ['customer_name','gender_guess','phone','mobile','email','wechat','company_name','department','job_title','city','province']:
        best,conflicts=pick_best(cands.get(field,[])); record[field]=best
        if conflicts:
            conflict_notes.append(f"{field} 存在候选冲突: {best} | {' | '.join(conflicts[:3])}")
    if not record['customer_name']:
        hint,_=pick_best(cands.get('remarks_name_hint',[]))
        if hint:
            conflict_notes.append(f'仅识别到昵称或弱姓名线索: {hint}')
    if record['company_name']:
        record['company_short_name']=company_short_name(record['company_name'])
    record['customer_type']=infer_customer_type(text_blob,record['company_name'])
    summary=[]
    if record['company_name'] or record['customer_name']:
        summary.append('客户身份信息已部分识别')
    if record['product_interest']:
        summary.append(f"关注 {record['product_interest']}")
    if record['use_case']:
        summary.append(f"场景: {record['use_case']}")
    if record['current_status']:
        summary.append(f"当前状态: {record['current_status']}")
    if record['budget_info']:
        summary.append(f"预算: {record['budget_info']}")
    if record['timeline_info']:
        summary.append(f"时间: {record['timeline_info']}")
    record['communication_summary']='；'.join(summary) if summary else '对话信息有限，暂无法形成完整客户画像。'
    if not record['mobile'] and not record['phone'] and not record['email'] and not record['wechat']:
        record['risk_flags'].append('联系方式缺失')
    if not record['company_name']:
        record['risk_flags'].append('公司信息缺失')
    if not record['budget_info']:
        record['risk_flags'].append('预算未明确')
    if not record['timeline_info']:
        record['risk_flags'].append('时间计划未明确')
    if not record['decision_maker']:
        record['risk_flags'].append('决策信息不足')
    if raw_text_length<120 or len(lines)<3:
        record['risk_flags'].append('信息不足')
    if group_note:
        conflict_notes.append(group_note)
    if conflict_notes:
        record['risk_flags'].append('存在字段冲突或拆分歧义')
    critical=['customer_name','mobile','email','wechat','company_name','product_interest','current_status']
    for field,value in record.items():
        if field in {'confidence','raw_text_length'}:
            continue
        if isinstance(value,list) and not value:
            record['missing_fields'].append(field)
        elif isinstance(value,str) and not value:
            record['missing_fields'].append(field)
    missing_critical=[f for f in critical if not record[f]]
    contact_strength=bool(record['mobile'] or record['phone'] or record['email'] or record['wechat'])
    record['opportunity_level']=infer_opportunity_level(record['current_status'],record['budget_info'],record['timeline_info'],record['decision_maker'],record['risk_flags'],contact_strength)
    record['next_step']=infer_next_step(record['current_status'],record['missing_fields'],record['risk_flags'])
    confidence=30
    if record['customer_name']:
        confidence+=10
    if record['company_name']:
        confidence+=12
    if contact_strength:
        confidence+=16
    if record['job_title']:
        confidence+=4
    if record['product_interest']:
        confidence+=8
    if record['use_case']:
        confidence+=6
    if record['current_status']!='初步接触':
        confidence+=8
    if record['budget_info']:
        confidence+=4
    if record['timeline_info']:
        confidence+=4
    confidence-=min(24,len(missing_critical)*3)
    confidence-=min(12,len(conflict_notes)*4)
    if '信息不足' in record['risk_flags']:
        confidence-=12
    if not customer_lines:
        confidence-=8
    record['confidence']=max(10,min(95,confidence))
    record['risk_flags']=normalize_list(record['risk_flags']+suspicious_flags(record))
    record['remarks']='；'.join(conflict_notes)
    return record

def extract_records_from_text(text,source_file,extraction_time):
    lines=parse_lines(normalize_text(text))
    if not lines:
        return [build_blank_record(source_file,extraction_time,len(text))]
    groups,group_note=split_groups(lines)
    return [extract_record(group,source_file,extraction_time,len(text),group_note if len(groups)==1 else '') for group in groups]

def path_is_text_candidate(path:Path):
    if path.name.startswith('customer_export_') or path.suffix.lower()=='.xlsx':
        return False
    return path.suffix.lower() in TEXT_SUFFIXES

def stringify_for_excel(value):
    if isinstance(value,list):
        return ' | '.join(str(x) for x in value)
    return '' if value is None else str(value)

def export_json(records,path:Path):
    path.write_text(json.dumps({'records':records},ensure_ascii=False,indent=2)+'\n',encoding='utf-8')

def export_markdown(records,path:Path):
    lines=['# 客户信息抽取摘要','',f'- 识别记录数：{len(records)}','']
    for idx,record in enumerate(records,start=1):
        lines.append(f'## 记录 {idx}')
        lines.append(f"- 姓名：{record['customer_name'] or '空'}")
        lines.append(f"- 公司：{record['company_name'] or '空'}")
        lines.append(f"- 电话/手机：{record['mobile'] or record['phone'] or '空'}")
        lines.append(f"- 邮箱：{record['email'] or '空'}")
        lines.append(f"- 当前状态：{record['current_status']}")
        lines.append(f"- 商机等级：{record['opportunity_level'] or '空'}")
        lines.append(f"- 下一步动作：{record['next_step']}")
        lines.append(f"- 缺失字段：{', '.join(record['missing_fields'][:10]) if record['missing_fields'] else '无'}")
        lines.append(f"- 歧义/备注：{record['remarks'] or '无'}")
        lines.append('')
    path.write_text('\n'.join(lines)+'\n',encoding='utf-8')

def export_xlsx(records,path:Path,field_config):
    wb=Workbook(); ws=wb.active; ws.title='customers'
    fields=field_config['field_order']; headers=[field_config['excel_headers'].get(f,f) for f in fields]
    ws.append(headers); fill=PatternFill(fill_type='solid',fgColor='D9EAF7'); font=Font(bold=True)
    for cell in ws[1]:
        cell.fill=fill; cell.font=font; cell.alignment=Alignment(vertical='top',wrap_text=True)
    for record in records:
        ws.append([stringify_for_excel(record.get(f,'')) for f in fields])
    for col in ws.columns:
        values=[stringify_for_excel(c.value) for c in col]; width=min(50,max(10,max(len(v) for v in values)+2))
        ws.column_dimensions[col[0].column_letter].width=width
        for cell in col:
            cell.alignment=Alignment(vertical='top',wrap_text=True)
    wb.save(path)

def default_output_path(input_path:Path,suffix:str,batch=False):
    return input_path / f'customer_export_batch.{suffix}' if batch else input_path.parent / f'customer_export_{input_path.stem}.{suffix}'

def process_single_file(path:Path,extraction_time:str):
    text,enc=read_text_file(path); records=extract_records_from_text(text,path.name,extraction_time)
    for record in records:
        record['remarks']=(record['remarks']+f'；检测编码: {enc}').strip('；')
    return records

def collect_input_files(path:Path):
    return [p for p in sorted(path.rglob('*')) if p.is_file() and path_is_text_candidate(p)]

def main():
    parser=argparse.ArgumentParser(description='Extract customer information from dialogue text and export JSON/XLSX/Markdown.')
    parser.add_argument('input_path',help='Input file path or directory path')
    parser.add_argument('--output-json',dest='output_json',help='Output JSON path')
    parser.add_argument('--output-xlsx',dest='output_xlsx',help='Output XLSX path')
    parser.add_argument('--output-md',dest='output_md',help='Output Markdown path')
    parser.add_argument('--batch',action='store_true',help='Treat input path as a directory and process all readable text files')
    args=parser.parse_args(); field_config=load_field_config(); input_path=Path(args.input_path).expanduser().resolve()
    if not input_path.exists():
        raise SystemExit(f'输入路径不存在: {input_path}')
    extraction_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if input_path.is_dir():
        if not args.batch:
            raise SystemExit('输入是目录时必须显式传入 --batch')
        records=[]
        for file_path in collect_input_files(input_path):
            records.extend(process_single_file(file_path,extraction_time))
        if not records:
            raise SystemExit('目录下没有可处理的文本文件')
        json_path=Path(args.output_json) if args.output_json else default_output_path(input_path,'json',True)
        xlsx_path=Path(args.output_xlsx) if args.output_xlsx else default_output_path(input_path,'xlsx',True)
        md_path=Path(args.output_md) if args.output_md else default_output_path(input_path,'md',True)
    else:
        records=process_single_file(input_path,extraction_time)
        json_path=Path(args.output_json) if args.output_json else default_output_path(input_path,'json')
        xlsx_path=Path(args.output_xlsx) if args.output_xlsx else default_output_path(input_path,'xlsx')
        md_path=Path(args.output_md) if args.output_md else default_output_path(input_path,'md')
    json_path.parent.mkdir(parents=True,exist_ok=True); xlsx_path.parent.mkdir(parents=True,exist_ok=True); md_path.parent.mkdir(parents=True,exist_ok=True)
    export_json(records,json_path); export_xlsx(records,xlsx_path,field_config); export_markdown(records,md_path)
    print(json.dumps({'record_count':len(records),'json_output':str(json_path),'xlsx_output':str(xlsx_path),'markdown_output':str(md_path)},ensure_ascii=False))
    return 0

if __name__=='__main__':
    sys.exit(main())
